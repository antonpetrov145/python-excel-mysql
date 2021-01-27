import openpyxl
import MySQLdb

db = MySQLdb.connect(host='localhost', user='root', db='testdb', passwd='Ant0n123')
cursor = db.cursor()

wb = openpyxl.load_workbook('ACTIVE_SIM_112018.xlsx')
ws = wb['активни СИМ_112018_детайлно']

rows = []

for row in ws.iter_rows(min_row=1, min_col=1, values_only=True):
    number = ws.cell(1,1).value
    chip = ws.cell(1,3).value
    rows.append((number, chip))

db = MySQLdb.connect(host='localhost', user='root', db='testdb', passwd='Ant0n123')
cursor = db.cursor()
cursor.executemany("""INSERT INTO sims (num,chip) VALUES (%s,%s)""", (rows))

db.commit()
db.close()


#import openpyxl
#import MySQLdb#


#wb = openpyxl.load_workbook('ACTIVE_SIM_112018.xlsx')
#ws = wb['активни СИМ_112018_детайлно']

# map is a convenient way to construct a list. you can get a 2x2 tuple by slicing 
# openpyxl.worksheet.worksheet.Worksheet instance and last row of worksheet 
# from openpyxl.worksheet.worksheet.Worksheet.max_row
#data = map(lambda x: {'num': x[0].value, 
#                      'chip': x[1].value}, 
#                ws[16: ws.max_row])

# filter is another builtin function. Filter blank cells out if needed
#data = filter(lambda x: None not in x.values(), data)

#db = MySQLdb.connect(host='localhost', user='root', db='testdb', passwd='Ant0n123')
#cursor = db.cursor()
#for row in data:
    # execute raw MySQL syntax by using execute function
#    cursor.execute('insert into table (num, chip)'
#                   'values ("{num}", "{chip}");'
#                   .format(**row))  # construct MySQL syntax through format function
#db.commit()